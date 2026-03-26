"""
Parse ipl_2026_squads.xlsx and seed the Player table.
Idempotent: skips seeding if players already exist.
"""
import os
import sys
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session


def seed_roster(db: Session) -> None:
    from models import Player

    if db.query(Player).count() > 0:
        return  # already seeded

    xlsx_path = Path(__file__).parent.parent / "ipl_2026_squads.xlsx"
    if not xlsx_path.exists():
        print(f"WARNING: {xlsx_path} not found — roster not seeded", file=sys.stderr)
        return

    wb = openpyxl.load_workbook(xlsx_path)
    players = []

    for team in wb.sheetnames:
        ws = wb[team]
        for row in ws.iter_rows(min_row=8, values_only=True):
            name, role, country = row[1], row[2], row[3]
            if name and role:
                players.append(
                    Player(
                        name=str(name).strip(),
                        team=team.strip(),
                        role=str(role).strip(),
                        country=str(country).strip() if country else "Unknown",
                    )
                )

    db.bulk_save_objects(players)
    db.commit()
    print(f"Seeded {len(players)} players from {xlsx_path.name}")
